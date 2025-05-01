import json
import csv
import openpyxl  # используется для load_xlsx()


def main():
    print("""Привет! Добро пожаловать в программу работы с банковскими транзакциями.
Выберите необходимый пункт меню:
1. Получить информацию о транзакциях из JSON-файла
2. Получить информацию о транзакциях из CSV-файла
3. Получить информацию о транзакциях из XLSX-файла""")

    while True:
        choice = input("Пользователь: ").strip()

        if choice in ("1", "2", "3"):
            break
        print("Неверный ввод. Пожалуйста, выберите 1, 2 или 3.")

    file_path = input("Введите путь к файлу: ")

    try:
        if choice == "1":
            transactions = load_json(file_path)
        elif choice == "2":
            transactions = load_csv(file_path)
        else:
            transactions = load_xlsx(file_path)

        display_transactions(transactions)

    except Exception as e:
        print(f"Ошибка: {e}")


def load_json(file_path):
    """Загрузка транзакций из JSON"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_csv(file_path):
    """Загрузка транзакций из CSV"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def load_xlsx(file_path):
    """Загрузка транзакций из Excel"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    transactions = []

    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        transactions.append(dict(zip(headers, row)))

    return transactions


def display_transactions(transactions):
    """Вывод транзакций в удобном формате"""
    print("\nСписок транзакций:")
    for i, tx in enumerate(transactions, 1):
        print(f"{i}. {tx.get('description', 'Без описания')} - {tx.get('amount', 0)} руб.")


if __name__ == "__main__":
    main()
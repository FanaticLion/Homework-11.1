import json
import csv
import openpyxl
from pathlib import Path

class FileHandler:
    @staticmethod
    def load_json(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    @staticmethod
    def load_csv(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return list(csv.DictReader(f))

    @staticmethod
    def load_xlsx(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

    @staticmethod
    def save_report(data, file_path):
        ext = Path(file_path).suffix.lower()
        if ext == '.json':
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        elif ext == '.csv':
            with open(file_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)